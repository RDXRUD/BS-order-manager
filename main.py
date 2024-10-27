import os
from openpyxl import load_workbook
import win32com.client

input_file="everest_.xlsx"
workbook = load_workbook(input_file)
# sheet = workbook.active

sheet_names = workbook.sheetnames 
print(sheet_names)

if len(sheet_names)>1:
    for sheet_name in sheet_names[1:]:
        del workbook[sheet_name]

sheet_names = workbook.sheetnames 
print(sheet_names)

sheet = workbook[sheet_names[0]]

find_qty = False
find_serial = False
qty_column = None
serial_column = None
serial_count=0
blank_line=0

file_name, file_extension = os.path.splitext(input_file)
output_file = f"{file_name}_modified{file_extension}"

for row in sheet.iter_rows():
    for cell in row:
        if cell.value is not None:
            if isinstance(cell.value, str) and cell.value.strip().lower() == "qty.":
                find_qty = True
                qty_column = cell.column 
                print(f"'qty' found at Row: {cell.row}, Column: {cell.column}")
                break
            
    if find_qty:
        break

if qty_column:
    rows_to_delete = []

    for row in sheet.iter_rows(min_row=cell.row + 1):
        qty_cell = row[qty_column - 1] 
        
        if row[1].value not in (None, '') or row[0].value not in (None, '') : 
            if not isinstance(qty_cell.value, int):
                if not isinstance(row[0].value,str):
                    print("row val:",row[0].value)
                    rows_to_delete.append(row[0].row) 
                
            else:
                if blank_line>0:
                    blank_line=0

                serial_count+=1
                row[0].value=serial_count
                
        else:
            blank_line+=1
            if blank_line>1:
                rows_to_delete.append(row[0].row)

    for row in reversed(rows_to_delete):
        sheet.delete_rows(row)
        
workbook.save(output_file)

def convert_excel_to_pdf(excel_file, pdf_file):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    workbook = excel.Workbooks.Open(excel_file)

    sheet = workbook.Sheets(1) 
    sheet_name = sheet.Name
    print(f"Converting sheet: {sheet_name}")

    try:
        sheet.Select()
        workbook.ExportAsFixedFormat(0, pdf_file)
        print(f"Successfully converted '{sheet_name}' to PDF.")
    except Exception as e:
        print(f"Error: {e}")
    finally:
        workbook.Close(False)
        excel.Quit()


excel_file = f"D:\PROJECTS\BS file manager\BS-order-manager\{file_name}_modified.xlsx"
pdf_file = f"D:\PROJECTS\BS file manager\BS-order-manager\{file_name}_modified.pdf"
convert_excel_to_pdf(excel_file, pdf_file)

